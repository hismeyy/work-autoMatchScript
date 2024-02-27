import os
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import time
from ttkbootstrap import ttk
from ttkbootstrap import Style
from openpyxl import load_workbook

file_one = ""
file_two = ""


def upload_reader_data():
    global file_one
    filepath = filedialog.askopenfilename()
    if filepath:
        entry_reader_data_path.config(state=tk.NORMAL)
        entry_reader_data_path.delete(0, tk.END)
        entry_reader_data_path.insert(0, filepath)
        entry_reader_data_path.config(state=tk.DISABLED)
        # 在这里添加处理上传文件的逻辑
        file_one = filepath


def upload_matching_data():
    global file_two
    filepath = filedialog.askopenfilename()
    if filepath:
        entry_matching_data_path.config(state=tk.NORMAL)
        entry_matching_data_path.delete(0, tk.END)
        entry_matching_data_path.insert(0, filepath)
        entry_matching_data_path.config(state=tk.DISABLED)
        # 在这里添加处理上传文件的逻辑
        file_two = filepath


def generate_files():
    if file_one == "" or file_two == "":
        messagebox.showinfo("提醒", "请先上传文件")
        return
    folderpath = filedialog.askdirectory()
    if folderpath:
        wb1 = load_workbook(filename=file_one)
        ws1 = wb1[wb1.sheetnames[0]]
        column_data = []
        for cell in ws1['A']:
            column_data.append(cell.value)

        wb2 = load_workbook(filename=file_two)
        ws2 = wb2[wb2.sheetnames[0]]

        max_row = 1
        for cell in ws2['A']:
            if cell.value is None:
                max_row = cell.row

        result = 0
        block = 100 / max_row

        i = 1
        ws2['D1'] = "Attendance"
        for cell in ws2['C']:
            if i > 1:
                if cell.value in column_data:
                    ws2['D' + str(i)] = "Present 出席"
                else:
                    ws2['D' + str(i)] = "Absent 缺席"
            i = i + 1
            result = result + block
            if result > 100:
                result = 100
            progress_var.set(result)
            app.update_idletasks()

        file_name = '结果数据.xlsx'
        file_path = os.path.join(folderpath, file_name)
        # 保存工作簿
        wb2.save(file_path)
        messagebox.showinfo("信息", "文件已成功生成于: " + folderpath)
        progress_var.set(0)
        app.update_idletasks()


app = tk.Tk()
app.withdraw()
style = Style(theme='sandstone')
TOP6 = style.master
app.title("自动匹配脚本")
app.iconbitmap('./logo.ico')
app.update_idletasks()
window_width = app.winfo_width()
window_height = app.winfo_height()
screen_width = app.winfo_screenwidth()
screen_height = app.winfo_screenheight()
x = (screen_width / 2) - (window_width / 2) - 170
y = (screen_height / 2) - (window_height / 2) - 100
app.geometry('+%d+%d' % (x, y))
app.resizable(width=False, height=False)
frame = tk.Frame(app)
frame.pack(padx=10, pady=10)
btn_upload_reader_data = tk.Button(frame, text="上传读卡器数据", command=upload_reader_data, width=15)
btn_upload_reader_data.grid(row=0, column=0, padx=10)
entry_reader_data_path = tk.Entry(frame, state=tk.DISABLED, width=60)
entry_reader_data_path.grid(row=0, column=1, pady=10, padx=10)
btn_upload_matching_data = tk.Button(frame, text="上传匹配数据", command=upload_matching_data, width=15)
btn_upload_matching_data.grid(row=1, column=0, pady=10)
entry_matching_data_path = tk.Entry(frame, state=tk.DISABLED, width=60)
entry_matching_data_path.grid(row=1, column=1, pady=10, padx=10)
btn_generate = tk.Button(frame, text="生成结果数据", command=generate_files, width=15)
btn_generate.grid(row=2, column=0, columnspan=2, pady=10)

progress_var = tk.DoubleVar()
progress_bar = ttk.Progressbar(frame, variable=progress_var, maximum=100)
progress_bar.grid(row=3, column=0, columnspan=2, pady=(0, 10), sticky='ew')

app.deiconify()
app.mainloop()
